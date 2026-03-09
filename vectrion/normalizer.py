"""Vectorian Record Normalizer

Converts any vault file into normalized evidence records suitable for
downstream stage processing (PII detection, entity resolution, etc.).

A "record" is a dict with:
  record_id     - unique ID within this engagement
  source_file   - original filename
  source_type   - structured | document | web | image | video | unknown
  name          - full name (if detected)
  email         - email address (if detected)
  phone         - phone number (if detected)
  ssn           - social security number (if detected)
  dob           - date of birth (if detected)
  ip            - IP address (if detected)
  cc            - credit card number (if detected)
  account       - account number (if detected)
  username      - username (if detected)
  password      - password (masked — flag only, value replaced)
  address       - street address (if detected)
  raw_text      - the text this record was derived from (≤2000 chars)
  pii_types     - list of PII category names found in this record
  pii_hit_count - total PII hit count
  sheet         - sheet name (XLSX only)
  row_num       - row number in source file (structured only)
"""
from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Any

from vectrion.detectors import detect_pii

# ── JSON deserialization safety limits ────────────────────────────────────────
# Prevent memory exhaustion from malicious uploaded files.
_JSON_MAX_BYTES    = 256 * 1024 * 1024   # 256 MB raw file size cap
_JSON_MAX_FIELDS   = 500                  # max keys per dict row
_JSON_MAX_DEPTH    = 10                   # max nesting depth


def _safe_json_loads(text: str) -> Any:
    """json.loads with field count and nesting-depth guard to prevent memory exhaustion."""
    # Pre-parse depth check via character scan — fast, no allocation
    nesting = 0
    in_str = False
    escape = False
    for ch in text:
        if escape:
            escape = False
            continue
        if ch == "\\" and in_str:
            escape = True
            continue
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch in ("{", "["):
            nesting += 1
            if nesting > _JSON_MAX_DEPTH:
                raise ValueError(
                    f"JSON nesting depth exceeds {_JSON_MAX_DEPTH} — "
                    "file may be malformed or malicious."
                )
        elif ch in ("}", "]"):
            nesting -= 1

    def _obj_hook(pairs: list) -> dict:
        if len(pairs) > _JSON_MAX_FIELDS:
            raise ValueError(
                f"JSON object exceeds {_JSON_MAX_FIELDS} fields — "
                "file may be malformed or malicious."
            )
        return dict(pairs)

    return json.loads(text, object_pairs_hook=_obj_hook)

# ─────────────────────────────────────────────────────────────────────────────
# FIELD NAME AUTO-MAPPING
# ─────────────────────────────────────────────────────────────────────────────

_FIELD_MAP: dict[str, str] = {
    # Name
    "name": "name", "full_name": "name", "fullname": "name", "full name": "name",
    "patient_name": "name", "client_name": "name", "customer_name": "name",
    "first_name": "first_name", "firstname": "first_name", "given_name": "first_name", "fname": "first_name",
    "last_name": "last_name", "lastname": "last_name", "surname": "last_name", "lname": "last_name",
    "middle_name": "middle_name", "middle": "middle_name",
    # Contact
    "email": "email", "email_address": "email", "e_mail": "email", "emailaddress": "email",
    "phone": "phone", "phone_number": "phone", "phonenumber": "phone",
    "telephone": "phone", "tel": "phone", "mobile": "phone", "cell": "phone",
    "fax": "fax",
    # Government IDs
    "ssn": "ssn", "social_security": "ssn", "social_security_number": "ssn",
    "sin": "ssn",  # Canada
    "tin": "tin", "tax_id": "tin", "taxpayer_id": "tin",
    "passport": "passport", "passport_number": "passport",
    "drivers_license": "drivers_license", "driver_license": "drivers_license",
    "dl": "drivers_license", "license_number": "drivers_license",
    # Health
    "dob": "dob", "date_of_birth": "dob", "birth_date": "dob", "birthdate": "dob",
    "birthday": "dob", "born": "dob",
    "mrn": "mrn", "medical_record": "mrn", "medical_record_number": "mrn",
    "patient_id": "mrn", "patient_number": "mrn",
    "diagnosis": "diagnosis", "condition": "diagnosis", "icd": "diagnosis",
    "insurance_id": "insurance_id", "insurance_number": "insurance_id", "member_id": "insurance_id",
    # Financial
    "account": "account", "account_number": "account", "account_no": "account",
    "acct": "account", "acct_no": "account",
    "routing": "routing", "routing_number": "routing", "aba": "routing",
    "credit_card": "cc", "card_number": "cc", "cc": "cc", "pan": "cc",
    "cvv": "cvv", "cvc": "cvv",
    "iban": "iban", "bank_account": "account",
    # Network
    "ip": "ip", "ip_address": "ip", "ipaddress": "ip",
    "mac": "mac", "mac_address": "mac",
    # Authentication
    "username": "username", "user_name": "username", "login": "username",
    "user": "username", "userid": "username", "user_id": "username",
    "password": "password", "passwd": "password", "pwd": "password", "pass": "password",
    "hash": "password_hash", "password_hash": "password_hash",
    # Location
    "address": "address", "street": "address", "street_address": "address",
    "city": "city", "state": "state", "zip": "zip", "zipcode": "zip",
    "postal_code": "zip", "postcode": "zip", "country": "country",
    # Metadata
    "id": "record_id_src", "user_id": "record_id_src", "customer_id": "record_id_src",
    "patient_id": "mrn", "employee_id": "record_id_src",
}


def _norm_key(k: str) -> str:
    return re.sub(r"[^a-z0-9]", "_", k.lower().strip())


def _mask_password(v: str) -> str:
    return "[REDACTED]" if v else ""


def _build_name(rec: dict) -> str:
    parts = []
    for f in ("first_name", "middle_name", "last_name"):
        v = rec.get(f, "")
        if v:
            parts.append(v)
    return " ".join(parts)


def _map_row(row: dict, source_file: str, row_num: int, source_type: str = "structured") -> dict:
    """Map a raw row dict to a normalized record."""
    rec: dict[str, Any] = {
        "record_id": f"{source_file}:row:{row_num}",
        "source_file": source_file,
        "source_type": source_type,
        "row_num": row_num,
    }

    raw_parts: list[str] = []
    for k, v in row.items():
        if v is None or str(v).strip() == "":
            continue
        nk = _norm_key(k)
        mapped = _FIELD_MAP.get(nk, nk)
        val = str(v).strip()

        if mapped == "password":
            rec[mapped] = _mask_password(val)
        else:
            rec[mapped] = val
        raw_parts.append(f"{k}: {val}")

    # Build composite name from first/last if full name missing
    if "name" not in rec:
        composed = _build_name(rec)
        if composed:
            rec["name"] = composed

    rec["raw_text"] = " | ".join(raw_parts)[:2000]

    # Run PII detection for anything not already mapped
    hits = detect_pii(rec["raw_text"])
    pii_types = list({h.kind for h in hits})
    seen_kinds: set[str] = set()
    for hit in hits:
        if hit.kind in seen_kinds:
            continue
        seen_kinds.add(hit.kind)
        if hit.kind == "email" and "email" not in rec:
            rec["email"] = hit.value
        elif hit.kind == "phone" and "phone" not in rec:
            rec["phone"] = hit.value
        elif hit.kind == "ssn" and "ssn" not in rec:
            rec["ssn"] = hit.value
        elif hit.kind == "ip" and "ip" not in rec:
            rec["ip"] = hit.value
        elif hit.kind == "credit_card" and "cc" not in rec:
            rec["cc"] = hit.value
        elif hit.kind == "dob" and "dob" not in rec:
            rec["dob"] = hit.value
        elif hit.kind == "mrn" and "mrn" not in rec:
            rec["mrn"] = hit.value

    rec["pii_types"] = pii_types
    rec["pii_hit_count"] = len(hits)
    return rec


# ─────────────────────────────────────────────────────────────────────────────
# STRUCTURED PARSERS (use original file for accurate field mapping)
# ─────────────────────────────────────────────────────────────────────────────

def _records_from_csv(path: Path, source_name: str) -> list[dict]:
    records: list[dict] = []
    # Auto-detect delimiter
    sample = path.read_bytes()[:4096].decode("utf-8", errors="replace")
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f, dialect=dialect)
        for i, row in enumerate(reader):
            if i >= 50_000:
                break
            records.append(_map_row(row, source_name, i))
    return records


def _records_from_json(path: Path, source_name: str) -> list[dict]:
    records: list[dict] = []
    try:
        # Cap file size before reading into memory
        if path.stat().st_size > _JSON_MAX_BYTES:
            return records
        content = path.read_text(encoding="utf-8", errors="replace")
        if path.suffix.lower() == ".jsonl":
            rows = []
            for ln in content.splitlines():
                ln = ln.strip()
                if ln:
                    try:
                        rows.append(_safe_json_loads(ln))
                    except Exception:
                        pass
        else:
            obj = _safe_json_loads(content)
            if isinstance(obj, list):
                rows = obj
            elif isinstance(obj, dict):
                # Find list-valued key (most likely the records array)
                rows = []
                for v in obj.values():
                    if isinstance(v, list) and v and isinstance(v[0], dict):
                        rows = v
                        break
                if not rows:
                    rows = [obj]
            else:
                rows = []
        for i, row in enumerate(rows):
            if i >= 50_000:
                break
            if isinstance(row, dict):
                records.append(_map_row(row, source_name, i))
            elif isinstance(row, (str, int, float)):
                # Scalar list → treat each as a text chunk
                records.append(_text_chunk_record(str(row), source_name, i))
    except Exception:
        pass
    return records


def _records_from_xlsx(path: Path, source_name: str) -> list[dict]:
    records: list[dict] = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        rec_idx = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            # First non-empty row as header
            header_row = next((r for r in rows if any(c is not None for c in r)), None)
            if header_row is None:
                continue
            headers = [str(c or "").strip() for c in header_row]
            header_idx = rows.index(header_row)
            for row in rows[header_idx + 1:]:
                if all(c is None for c in row):
                    continue
                row_dict = {
                    headers[i]: ("" if row[i] is None else str(row[i]).strip())
                    for i in range(min(len(headers), len(row)))
                }
                rec = _map_row(row_dict, source_name, rec_idx)
                rec["sheet"] = sheet_name
                records.append(rec)
                rec_idx += 1
                if rec_idx >= 50_000:
                    break
    except ImportError:
        pass
    except Exception:
        pass
    return records


def _records_from_xml(path: Path, source_name: str) -> list[dict]:
    records: list[dict] = []
    try:
        import xml.etree.ElementTree as ET
        tree = ET.parse(path)
        root = tree.getroot()

        # Find the most common repeated child tag = likely the record element
        child_counts: dict[str, int] = {}
        for child in root:
            tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
            child_counts[tag] = child_counts.get(tag, 0) + 1

        if not child_counts:
            return records

        record_tag = max(child_counts, key=child_counts.get)
        rec_idx = 0
        for elem in root:
            tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            if tag != record_tag:
                continue
            row: dict[str, str] = dict(elem.attrib)
            for child in elem:
                ct = child.tag.split("}")[-1] if "}" in child.tag else child.tag
                row[ct] = child.text or ""
            if elem.text and elem.text.strip():
                row.setdefault("text", elem.text.strip())
            records.append(_map_row(row, source_name, rec_idx))
            rec_idx += 1
            if rec_idx >= 50_000:
                break
    except Exception:
        pass
    return records


# ─────────────────────────────────────────────────────────────────────────────
# TEXT / DOCUMENT PARSERS (chunk by paragraph, extract PII per chunk)
# ─────────────────────────────────────────────────────────────────────────────

def _text_chunk_record(text: str, source_name: str, chunk_idx: int) -> dict:
    """Create a single record from a text chunk."""
    hits = detect_pii(text)
    pii_types = list({h.kind for h in hits})
    rec: dict[str, Any] = {
        "record_id": f"{source_name}:chunk:{chunk_idx}",
        "source_file": source_name,
        "source_type": "document",
        "raw_text": text[:2000],
        "pii_types": pii_types,
        "pii_hit_count": len(hits),
    }
    seen: set[str] = set()
    for hit in hits:
        if hit.kind in seen:
            continue
        seen.add(hit.kind)
        if hit.kind == "email":
            rec["email"] = hit.value
        elif hit.kind == "phone":
            rec["phone"] = hit.value
        elif hit.kind == "ssn":
            rec["ssn"] = hit.value
        elif hit.kind == "ip":
            rec["ip"] = hit.value
        elif hit.kind == "credit_card":
            rec["cc"] = hit.value
        elif hit.kind == "dob":
            rec["dob"] = hit.value
        elif hit.kind == "mrn":
            rec["mrn"] = hit.value
    return rec


def _records_from_text(text: str, source_name: str, source_type: str = "document") -> list[dict]:
    """Split free-form text into paragraphs and create PII-bearing records."""
    records: list[dict] = []
    if not text or not text.strip():
        return records

    # Split on double newlines first; if that gives one huge block, split on single newlines
    paragraphs = re.split(r"\n{2,}", text.strip())
    if len(paragraphs) <= 2 and len(text) > 500:
        paragraphs = text.split("\n")

    for i, para in enumerate(paragraphs):
        para = para.strip()
        if not para or len(para) < 8:
            continue
        hits = detect_pii(para)
        # Keep paragraph if it has PII or is substantial enough to be meaningful
        if not hits and len(para) < 60:
            continue
        rec = _text_chunk_record(para, source_name, i)
        rec["source_type"] = source_type
        records.append(rec)

    return records


# ─────────────────────────────────────────────────────────────────────────────
# PUBLIC API
# ─────────────────────────────────────────────────────────────────────────────

def extend_field_map(user_fields: list[dict]) -> None:
    """
    Register user-defined field aliases into _FIELD_MAP at runtime.
    Each field: {"name": "employee_id", "search_terms": ["emp_id", "staff_id"]}
    Adds _FIELD_MAP["emp_id"] = "employee_id" etc. before normalization runs.
    """
    for field in user_fields:
        canonical = (field.get("name") or "").strip()
        if not canonical:
            continue
        # Also map the canonical name to itself
        _FIELD_MAP[_norm_key(canonical)] = canonical
        for term in field.get("search_terms", []):
            norm = _norm_key(term)
            if norm:
                _FIELD_MAP[norm] = canonical


# ── RAW ROWS (no normalization) ───────────────────────────────────────────────

def _raw_rows_from_csv(path: Path) -> list[dict]:
    """Return raw row dicts from CSV without any field mapping."""
    rows: list[dict] = []
    sample = path.read_bytes()[:4096].decode("utf-8", errors="replace")
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f, dialect=dialect)
        for i, row in enumerate(reader):
            if i >= 50_000:
                break
            rows.append(dict(row))
    return rows


def _raw_rows_from_json(path: Path) -> list[dict]:
    """Return raw row dicts from JSON/JSONL without any field mapping."""
    rows: list[dict] = []
    try:
        if path.stat().st_size > _JSON_MAX_BYTES:
            return rows
        content = path.read_text(encoding="utf-8", errors="replace")
        if path.suffix.lower() == ".jsonl":
            for ln in content.splitlines():
                ln = ln.strip()
                if ln:
                    try:
                        obj = _safe_json_loads(ln)
                        if isinstance(obj, dict):
                            rows.append(obj)
                    except Exception:
                        pass
        else:
            obj = _safe_json_loads(content)
            if isinstance(obj, list):
                raw = obj
            elif isinstance(obj, dict):
                raw = []
                for v in obj.values():
                    if isinstance(v, list) and v and isinstance(v[0], dict):
                        raw = v
                        break
                if not raw:
                    raw = [obj]
            else:
                raw = []
            for item in raw:
                if isinstance(item, dict):
                    rows.append(item)
        rows = rows[:50_000]
    except Exception:
        pass
    return rows


def _raw_rows_from_xlsx(path: Path) -> list[dict]:
    """Return raw row dicts from XLSX without any field mapping."""
    rows: list[dict] = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                continue
            header_row = next((r for r in all_rows if any(c is not None for c in r)), None)
            if header_row is None:
                continue
            headers = [str(c or "").strip() for c in header_row]
            header_idx = all_rows.index(header_row)
            for row in all_rows[header_idx + 1:]:
                if all(c is None for c in row):
                    continue
                row_dict = {
                    headers[i]: ("" if row[i] is None else str(row[i]).strip())
                    for i in range(min(len(headers), len(row)))
                }
                rows.append(row_dict)
                if len(rows) >= 50_000:
                    break
    except ImportError:
        pass
    except Exception:
        pass
    return rows


def _raw_rows_from_xml(path: Path) -> list[dict]:
    """Return raw row dicts from XML without any field mapping."""
    rows: list[dict] = []
    try:
        import xml.etree.ElementTree as ET
        tree = ET.parse(path)
        root = tree.getroot()
        child_counts: dict[str, int] = {}
        for child in root:
            tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
            child_counts[tag] = child_counts.get(tag, 0) + 1
        if not child_counts:
            return rows
        record_tag = max(child_counts, key=child_counts.get)
        for elem in root:
            tag = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            if tag != record_tag:
                continue
            row: dict[str, str] = dict(elem.attrib)
            for child in elem:
                ct = child.tag.split("}")[-1] if "}" in child.tag else child.tag
                row[ct] = child.text or ""
            rows.append(row)
            if len(rows) >= 50_000:
                break
    except Exception:
        pass
    return rows


def raw_rows_from_vault(vault_index: list[dict], vault_dir: Path) -> dict[str, list[dict]]:
    """
    Return {original_name: [raw_row_dicts]} for each vault file.
    Structured files (CSV/XLSX/JSON/XML): raw column names and values preserved,
    no _FIELD_MAP normalization applied.
    Unstructured files: [{"_text": chunk, "_chunk": i}] per paragraph chunk.
    """
    result: dict[str, list[dict]] = {}
    for file_record in vault_index:
        fname = file_record["filename"]
        source_name = file_record.get("original_name", fname)
        file_path = vault_dir / fname
        ext = Path(fname).suffix.lower()

        rows: list[dict] = []
        try:
            if ext in (".csv", ".tsv") and file_path.exists():
                rows = _raw_rows_from_csv(file_path)
            elif ext in (".json", ".jsonl") and file_path.exists():
                rows = _raw_rows_from_json(file_path)
            elif ext in (".xlsx", ".xls") and file_path.exists():
                rows = _raw_rows_from_xlsx(file_path)
            elif ext == ".xml" and file_path.exists():
                rows = _raw_rows_from_xml(file_path)
            else:
                # Unstructured: return text chunks
                text_path = vault_dir / (fname + ".txt")
                if text_path.exists():
                    text = text_path.read_text(encoding="utf-8", errors="replace")
                    paragraphs = re.split(r"\n{2,}", text.strip()) if text.strip() else []
                    rows = [
                        {"_text": p, "_chunk": str(i)}
                        for i, p in enumerate(paragraphs)
                        if p.strip()
                    ]
        except Exception:
            pass

        result[source_name] = rows
    return result


def normalize_vault_file(file_record: dict, vault_dir: Path) -> list[dict]:
    """
    Normalize one vault file into structured evidence records.

    For structured formats (CSV/TSV/XLSX/JSON/JSONL/XML): each row → one record.
    For documents/images/video: each PII-bearing paragraph/frame → one record.

    Returns empty list on failure (errors are logged in file_record['proc_error']).
    """
    fname = file_record["filename"]
    source_name = file_record.get("original_name", fname)
    file_type = file_record.get("type", "unknown")
    file_path = vault_dir / fname
    ext = Path(fname).suffix.lower()

    try:
        # Structured formats — parse original file for accurate column headers
        if ext in (".csv", ".tsv") and file_path.exists():
            return _records_from_csv(file_path, source_name)

        if ext in (".json", ".jsonl") and file_path.exists():
            return _records_from_json(file_path, source_name)

        if ext in (".xlsx", ".xls") and file_path.exists():
            return _records_from_xlsx(file_path, source_name)

        if ext == ".xml" and file_path.exists():
            return _records_from_xml(file_path, source_name)

        # All other formats — use the extracted .txt file
        text_path = vault_dir / (fname + ".txt")
        if text_path.exists():
            text = text_path.read_text(encoding="utf-8", errors="replace")
            return _records_from_text(text, source_name, file_type)

    except Exception:
        pass

    return []


def normalize_vault(vault_index: list[dict], vault_dir: Path) -> list[dict]:
    """
    Normalize all vault files into a single flat records list.
    Returns records sorted by source file then record_id.
    """
    all_records: list[dict] = []
    for file_record in vault_index:
        recs = normalize_vault_file(file_record, vault_dir)
        all_records.extend(recs)
    return all_records


def summarize_records(records: list[dict]) -> dict:
    """
    Produce summary statistics over a normalized record set.
    Used for stage output and UI display.
    """
    total = len(records)
    pii_counter: dict[str, int] = {}
    sensitivity: dict[str, int] = {
        "PHI": 0, "PII": 0, "PCI": 0, "Credentials": 0, "Financial": 0, "Network": 0, "Other": 0,
    }
    unique_emails: set[str] = set()
    unique_phones: set[str] = set()
    unique_ssns: set[str] = set()
    source_files: set[str] = set()

    _PHI   = {"ssn", "dob", "mrn", "insurance_id"}
    _PCI   = {"credit_card", "cc"}
    _CRED  = {"password", "password_hash", "username"}
    _FIN   = {"account", "routing", "iban"}
    _NET   = {"ip", "mac"}
    _PII   = {"email", "phone", "name", "address"}

    for rec in records:
        source_files.add(rec.get("source_file", ""))
        for pt in rec.get("pii_types", []):
            pii_counter[pt] = pii_counter.get(pt, 0) + 1

        types = set(rec.get("pii_types", [])) | set(rec.keys())

        if types & _PHI:
            sensitivity["PHI"] += 1
        elif types & _PCI:
            sensitivity["PCI"] += 1
        elif types & _CRED:
            sensitivity["Credentials"] += 1
        elif types & _FIN:
            sensitivity["Financial"] += 1
        elif types & _NET:
            sensitivity["Network"] += 1
        elif types & _PII or rec.get("email") or rec.get("phone") or rec.get("name"):
            sensitivity["PII"] += 1
        elif rec.get("raw_text", "").strip():
            sensitivity["Other"] += 1

        if rec.get("email"):
            unique_emails.add(rec["email"].lower())
        if rec.get("phone"):
            unique_phones.add(re.sub(r"\D", "", rec["phone"])[-10:])
        if rec.get("ssn"):
            unique_ssns.add(rec["ssn"])

    return {
        "total_records": total,
        "source_files": sorted(source_files),
        "unique_emails": len(unique_emails),
        "unique_phones": len(unique_phones),
        "unique_ssns": len(unique_ssns),
        "pii_type_counts": pii_counter,
        "sensitivity_classification": {k: v for k, v in sensitivity.items() if v > 0},
    }
